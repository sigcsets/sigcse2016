from __future__ import unicode_literals

from openpyxl import load_workbook

import codecs
import re
import sys

#UTF8Reader = codecs.getreader('utf8')
#sys.stdin = UTF8Reader(sys.stdin)
#UTF8Writer = codecs.getwriter('utf8')
#sys.stdout = UTF8Writer(sys.stdout)

wb = load_workbook('workshops.xlsx')
# Get the first sheet in this workbook
ws = wb.get_sheet_by_name("Sheet1")

class Col:
    number, id, title, subtype, tags, abstract = range (0, 6)
    # Authors start at G, and go out to BF

def uni (value):
    value = re.sub(u"(\u2018|\u2019)", "'", value)
    value = re.sub(u"(\u2013)", "--", value)
    value = re.sub(u"(\u2014)", "---", value)
    value = re.sub('"', '\\"', value)
    value.encode('utf-8')

    return value

# Open a YAML file for Jekyll processing
of = open ('workshops.yaml', 'w')

for ndx, row in enumerate(ws.iter_rows('A2:F36')):
    rowndx = ndx + 2
    # The row is a list of Cell objects
    # print row[Col.number].value, row[Col.title].value
    authors = ""
    auth_range = list(ws['G{0}'.format(rowndx):'BF{0}'.format(rowndx)])[0]
    # print auth_range
    # name, org, cnty, email
    #   0    1    2     3
    content = auth_range[0].value
    # print ("CONTENT: {0}".format(content))
    offset = 0

    while content is not None:
        # print "OFFSET: %s" % offset
        authors += \
'''
    - name: "{0}"
      org: "{1}"
      country: "{2}"
      email: "{3}"
'''.format( uni(auth_range[0+offset].value),
            uni(auth_range[1+offset].value),
            uni(auth_range[2+offset].value),
            uni(auth_range[3+offset].value))
        offset += 4
        # print authors
        content = list(auth_range)[offset].value
    # print "ABSTRACT: %s " % Col.abstract
    block = \
'''
- number: {0}
  id: {1}
  title: "{2}"
  tags: "{3}"
  abstract: "{4}"
  authors: {5}
 '''.format(row[Col.number].value,
            row[Col.id].value,
            uni(row[Col.title].value),
            uni(row[Col.tags].value),
            uni(row[Col.abstract].value),
            authors
            )

    print block.encode('utf-8')
    of.write(block.encode('utf-8'))

of.close()
